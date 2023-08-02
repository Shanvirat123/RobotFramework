import openpyxl

def read_excel_data(file_path, sheet_name, row_number, column_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    # Get the column index based on the column name
    column_index = None
    for col_index, cell in enumerate(sheet[1], 1):
        if cell.value == column_name:
            column_index = col_index

    # Read data from the specified cell
    if column_index and 1 <= int(row_number) <= sheet.max_row:
        data = sheet.cell(row=int(row_number), column=column_index).value
    else:
        data = None

    workbook.close()
    return data
