# ExcelLibrary.py

import openpyxl

def read_excel_data(file_path, sheet_name):
    data = []
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if all(row):  # Check if all cells in the row are not empty
            data.append({sheet.cell(1, col_index).value: cell_value for col_index, cell_value in enumerate(row, 1)})

    workbook.close()
    return data
